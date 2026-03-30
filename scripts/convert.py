"""Excel → JSON conversion for apartment data."""
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def is_low_floor(floor_str: str) -> bool:
    """5층 이하 또는 '저층' 포함이면 True."""
    if "저층" in floor_str:
        return True
    try:
        current = int(floor_str.split("/")[0])
        return current <= 5
    except (ValueError, IndexError):
        return False


def format_price_range(prices: list[float]) -> str | None:
    """비저층 매물 가격으로 범위 문자열 생성. 0건이면 None."""
    if not prices:
        return None
    if len(prices) == 1:
        return f"{prices[0]:.1f}억"
    return f"{min(prices):.1f}~{max(prices):.1f}억"


def load_and_convert(xlsx_path: Path) -> dict:
    """Excel 파일을 읽어 JSON 구조로 변환."""
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    records = [dict(zip(headers, row)) for row in rows]

    # Filter: 300세대 이상, 9.5~11.5억
    filtered = [
        r for r in records
        if (r.get("세대수") or 0) >= 300
        and 9.5 <= (r.get("매매가(억)") or 0) <= 11.5
    ]

    # Group by 단지명 + 주소 + 면적타입
    groups: dict[tuple, list] = {}
    for r in filtered:
        key = (r["단지명"], r["주소"], r["면적타입"])
        groups.setdefault(key, []).append(r)

    apartments = []
    for (name, address, apt_type), listings in groups.items():
        first = listings[0]

        processed_listings = []
        non_low_prices = []

        for l in listings:
            floor_str = str(l.get("층수", ""))
            low = is_low_floor(floor_str)
            price = float(l["매매가(억)"])

            if not low:
                non_low_prices.append(price)

            processed_listings.append({
                "floor": floor_str,
                "price": price,
                "memo": l.get("메모") or "",
                "naver_url": l.get("네이버_URL") or "",
                "is_low_floor": low,
            })

        price_range = format_price_range(non_low_prices)
        if price_range is None:
            continue  # 비저층 매물 0건 → 제외

        apt_id = f"{name.replace(' ', '')}_{apt_type}"

        apartments.append({
            "id": apt_id,
            "name": name,
            "region": first.get("지역") or "",
            "address": address or "",
            "type": apt_type,
            "area_sqm": float(first.get("전용면적(㎡)") or 0),
            "built_year": int(first.get("입주년도") or 0),
            "total_units": int(first.get("세대수") or 0),
            "gangnam_minutes": int(first.get("강남역_시간(분)") or 0),
            "gangnam_transport": first.get("강남역_교통수단") or "",
            "nearest_station": first.get("최근역") or "",
            "station_walk_min": int(first.get("최근역_도보(분)") or 0),
            "mart": first.get("대형마트") or "",
            "commercial": first.get("상권") or "",
            "nimby": first.get("혐오시설") or "",
            "schools": first.get("주변학교") or "",
            "map_image": first.get("지도_이미지"),
            "price_range": price_range,
            "listings": processed_listings,
        })

    return {
        "generated_at": datetime.now().isoformat(timespec="minutes"),
        "apartments": apartments,
    }


def main():
    project_root = Path(__file__).resolve().parent.parent
    xlsx_path = project_root / "data" / "apartments.xlsx"
    web_dir = project_root / "web"
    maps_src = project_root / "data" / "maps"
    maps_dst = web_dir / "maps"

    if not xlsx_path.exists():
        print(f"Error: {xlsx_path} not found")
        sys.exit(1)

    result = load_and_convert(xlsx_path)

    # Write JSON
    web_dir.mkdir(exist_ok=True)
    output_path = web_dir / "apartments.json"
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2))
    print(f"Written {len(result['apartments'])} apartments to {output_path}")

    # Copy map images
    if maps_src.exists():
        maps_dst.mkdir(exist_ok=True)
        for img in maps_src.glob("*.png"):
            shutil.copy2(img, maps_dst / img.name)
        for img in maps_src.glob("*.jpg"):
            shutil.copy2(img, maps_dst / img.name)
        print(f"Copied map images to {maps_dst}")


if __name__ == "__main__":
    main()
